import os
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)




def isXLSX(filename):
#checks if file is in xlsx format
    split_file = filename.split(".")
    if(len(split_file) == 1 or split_file[len(split_file)-1]!='xlsx'):
        print(len(split_file))
        return False
    return True


def copy_cell_style(source_cell, target_cell):
    target_cell.font = source_cell.font.copy()
    target_cell.alignment = source_cell.alignment.copy()
    target_cell.border = source_cell.border.copy()
    target_cell.fill = source_cell.fill.copy()

def split_sheets(input_file,dir_store):
    #load workbook
    wb = openpyxl.load_workbook(input_file)

    #create the new excels for each sheet
    for sheet_name in wb.sheetnames:
        #create a new workbook
        new_wb = Workbook()
        new_ws = new_wb.active

        #get the original sheet
        original_ws = wb[sheet_name]

        #copies data from original
        for row in original_ws.iter_rows(values_only=True):
            new_ws.append(row)

        #copies the styling
        for row in original_ws.iter_rows():
            for cell in row:
                copy_cell_style(cell,new_ws[cell.coordinate])

        #save the new workbook
        target_name = sheet_name.strip()+".xlsx"
        target_file = os.path.join(dir_store,target_name)
        new_wb.save(target_file)


#creates the store directory on current dir
def create_directory_on_current_dir(sub_dir,curr_dir):
    try:
        #here the created files will be stored
        current_dir_path = os.path.join(curr_dir,sub_dir)
        
        #check if dir exists
        if not os.path.exists(current_dir_path):
            #create the dir
            os.makedirs(current_dir_path)
            print(f"Directory '{current_dir_path}' created successfully.")
        else:
            print(f"Directory '{current_dir_path}' already exists.")
        
        return current_dir_path
    except Exception as e:
        print(f"An error occurred while creating the directory: {e}")
        return None



def main():

    #source_file name taken from arguments
    while True:
        source_excel = input("Give filename(for example hello.xlsx): ")

        if(not isXLSX(source_excel)):
            print("File is not in xlsx format...try again")
            continue

        directory_name = source_excel.rsplit(".",1)[0] #splits on the last . only
        print(directory_name)
        #get the current directory
        current_path = os.path.abspath(sys.argv[0])
        current_dir = os.path.dirname(current_path)
        source_excel = os.path.join(current_dir, source_excel)
        print("\n"+source_excel)

        if(os.path.exists(source_excel)):
            break
        print("Wrong filename...try again")
    

    #the folder name that the excels will be stored
    storeDirectory = directory_name 



    #create the folder on the current dir
    store_dir = create_directory_on_current_dir(storeDirectory,current_dir)
    

    #now I have all the info I need to create the excel files
    split_sheets(source_excel,store_dir)
    print("Succesfully splitted the files!")

main()


##THE FILE MUST BE IN THE SAME DIR AS THE .exe##

##CANNOT COPY DATA VALIDATIONS##